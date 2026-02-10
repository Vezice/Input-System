"""
iBot v2 File Parser

Parses Excel (.xlsx, .xls) and CSV files for data import.
Handles Indonesian number formats and date formats.
Includes column mapping to normalize data to standard column order.
"""

import csv
import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from config import CategoryConfig
from column_mapper import ColumnMapper, create_mapper_for_category
from utils.logger import get_logger

logger = get_logger(__name__)

# Try to import openpyxl for Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger.warning("openpyxl not installed - Excel support disabled")

# Try to import xlrd for legacy .xls files
try:
    import xlrd
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False
    logger.warning("xlrd not installed - Legacy .xls support disabled")


class ParsedFile:
    """Result of parsing a file."""

    def __init__(
        self,
        filename: str,
        headers: List[str],
        rows: List[Dict[str, Any]],
        category: Optional[CategoryConfig],
        total_rows: int,
        skipped_rows: int = 0,
        errors: List[str] = None,
        column_mapped: bool = False,
        source_headers: List[str] = None,
    ):
        self.filename = filename
        self.headers = headers
        self.rows = rows
        self.category = category
        self.total_rows = total_rows
        self.skipped_rows = skipped_rows
        self.errors = errors or []
        self.column_mapped = column_mapped  # True if column mapping was applied
        self.source_headers = source_headers or []  # Original headers before mapping

    @property
    def is_valid(self) -> bool:
        return self.category is not None and len(self.rows) > 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "category": self.category.name if self.category else None,
            "headers": self.headers,
            "total_rows": self.total_rows,
            "parsed_rows": len(self.rows),
            "skipped_rows": self.skipped_rows,
            "errors": self.errors,
            "column_mapped": self.column_mapped,
        }


def clean_header(header: str) -> str:
    """
    Clean a header string for use as a column name.

    - Strips whitespace
    - Replaces spaces with underscores
    - Removes special characters
    - Converts to lowercase
    """
    if not header:
        return ""

    cleaned = str(header).strip().lower()

    # Replace spaces and special chars with underscores
    cleaned = re.sub(r'[\s\-\.]+', '_', cleaned)

    # Remove characters that aren't alphanumeric or underscore
    cleaned = re.sub(r'[^a-z0-9_]', '', cleaned)

    # Remove multiple consecutive underscores
    cleaned = re.sub(r'_+', '_', cleaned)

    # Remove leading/trailing underscores
    cleaned = cleaned.strip('_')

    return cleaned


def parse_number(value: Any) -> Optional[Union[int, float]]:
    """
    Parse numeric values, handling Indonesian number format.

    Indonesian format: 1.234.567,89 (dots for thousands, comma for decimal)
    Standard format: 1,234,567.89 (commas for thousands, dot for decimal)
    """
    if value is None:
        return None

    # Already numeric
    if isinstance(value, (int, float)):
        return value

    value_str = str(value).strip()
    if not value_str:
        return None

    # Treat "-" (dash) as null/missing value
    if value_str in ('-', '—', '–', 'N/A', 'n/a', '#N/A'):
        return None

    # Remove currency symbols and whitespace
    value_str = re.sub(r'[Rp\s\$€£¥]', '', value_str)

    # Handle percentage
    is_percentage = '%' in value_str
    value_str = value_str.replace('%', '')

    # Detect and handle Indonesian number format
    dot_count = value_str.count('.')
    comma_count = value_str.count(',')

    if dot_count > 1 or (dot_count == 1 and comma_count == 1 and value_str.index('.') < value_str.index(',')):
        # Indonesian format: convert to standard
        value_str = value_str.replace('.', '')  # Remove thousand separators
        value_str = value_str.replace(',', '.')  # Convert decimal separator
    elif comma_count > 1 or (comma_count == 1 and dot_count == 0):
        # Could be Indonesian with no decimal, or standard thousands
        if comma_count == 1 and len(value_str.split(',')[1]) == 3:
            value_str = value_str.replace(',', '')  # Remove thousand separator
        elif comma_count >= 1:
            value_str = value_str.replace(',', '.')

    # Now parse as standard format
    try:
        result = Decimal(value_str)
        if is_percentage:
            result = result / 100
        # Convert to int if no decimal places
        if result == int(result):
            return int(result)
        return float(result)
    except InvalidOperation:
        return None


def parse_cell_value(value: Any, column_name: str) -> Any:
    """
    Parse a cell value. Returns as string for BigQuery (all STRING columns).
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return ""

    # Return as string (BigQuery will store as STRING)
    # Keep "-" and other values as-is
    return str(value).strip()


def read_csv_file(
    file_content: Union[str, bytes],
    encoding: str = 'utf-8',
    delimiter: str = ',',
) -> Tuple[List[str], List[List[Any]]]:
    """
    Read CSV file content and return headers and rows.
    """
    if isinstance(file_content, bytes):
        # Try to decode, handling common encodings
        for enc in [encoding, 'utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                file_content = file_content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode CSV file with any known encoding")

    reader = csv.reader(io.StringIO(file_content), delimiter=delimiter)
    all_rows = list(reader)

    if not all_rows:
        return [], []

    headers = [str(h).strip() for h in all_rows[0]]
    data_rows = all_rows[1:]

    return headers, data_rows


def read_xlsx_file(
    file_content: bytes,
    sheet_name: Optional[str] = None,
    header_row: int = 0,
    data_start_row: int = 1,
) -> Tuple[List[str], List[List[Any]]]:
    """
    Read Excel .xlsx file and return headers and rows.
    """
    if not EXCEL_SUPPORT:
        raise ImportError("openpyxl is required for Excel support")

    # Note: read_only=True can cause issues with some files, especially small ones
    workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    all_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not all_rows or len(all_rows) <= header_row:
        logger.warning(f"Not enough rows in xlsx: got {len(all_rows)}, need > {header_row}")
        return [], []

    # Extract headers
    header_data = all_rows[header_row]
    headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(header_data)]

    # Extract data rows
    data_rows = all_rows[data_start_row:]

    return headers, data_rows


def read_xls_file(
    file_content: bytes,
    sheet_index: int = 0,
    header_row: int = 0,
    data_start_row: int = 1,
) -> Tuple[List[str], List[List[Any]]]:
    """
    Read legacy Excel .xls file and return headers and rows.
    """
    if not XLS_SUPPORT:
        raise ImportError("xlrd is required for .xls support")

    workbook = xlrd.open_workbook(file_contents=file_content)
    sheet = workbook.sheet_by_index(sheet_index)

    if sheet.nrows <= header_row:
        return [], []

    # Extract headers
    header_data = sheet.row_values(header_row)
    headers = [str(h).strip() if h else f"column_{i}" for i, h in enumerate(header_data)]

    # Extract data rows
    data_rows = []
    for row_idx in range(data_start_row, sheet.nrows):
        data_rows.append(sheet.row_values(row_idx))

    return headers, data_rows


def parse_file(
    filename: str,
    file_content: bytes,
    category: CategoryConfig,
) -> ParsedFile:
    """
    Parse a file and return structured data.

    Args:
        filename: Name of the file
        file_content: File content as bytes
        category: Category config (from folder path)

    Returns:
        ParsedFile with parsed data
    """
    logger.info(f"Parsing file: {filename}", category=category.name)

    errors = []
    headers = []
    raw_rows = []

    # Determine file type and read
    filename_lower = filename.lower()

    try:
        if filename_lower.endswith('.csv'):
            headers, raw_rows = read_csv_file(file_content)
        elif filename_lower.endswith('.xlsx'):
            headers, raw_rows = read_xlsx_file(
                file_content,
                header_row=category.header_row,
                data_start_row=category.data_start_row
            )
        elif filename_lower.endswith('.xls'):
            headers, raw_rows = read_xls_file(
                file_content,
                header_row=category.header_row,
                data_start_row=category.data_start_row
            )
        else:
            # Try to guess - attempt xlsx first, then csv
            try:
                headers, raw_rows = read_xlsx_file(
                    file_content,
                    header_row=category.header_row,
                    data_start_row=category.data_start_row
                )
            except Exception:
                try:
                    headers, raw_rows = read_csv_file(file_content)
                except Exception as e:
                    raise ValueError(f"Unsupported file format: {filename}") from e

    except Exception as e:
        logger.error(f"Error reading file {filename}: {e}")
        return ParsedFile(
            filename=filename,
            headers=[],
            rows=[],
            category=category,
            total_rows=0,
            errors=[str(e)],
        )

    if not headers:
        logger.error(f"No headers found in file: {filename}")
        return ParsedFile(
            filename=filename,
            headers=[],
            rows=[],
            category=category,
            total_rows=0,
            errors=["No headers found in file"],
        )

    # Filter out empty rows before mapping
    filtered_rows = []
    skipped = 0
    for row in raw_rows:
        if not row or all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            skipped += 1
            continue
        filtered_rows.append(row)

    # Apply column mapping to normalize data to standard column order
    # This handles files with different column layouts
    column_mapped = False
    source_headers = list(headers)  # Keep original headers

    if category.required_headers:
        try:
            mapper = create_mapper_for_category(
                category_name=category.name,
                standard_headers=category.required_headers,
                column_aliases=category.column_aliases,
            )

            # Map columns - this normalizes the data to standard column order
            mapped_headers, mapped_rows = mapper.map_file(headers, filtered_rows)

            logger.info(
                f"Column mapping applied: {len(headers)} source cols -> {len(mapped_headers)} standard cols",
                source_cols=len(headers),
                mapped_cols=len(mapped_headers),
            )

            # Use mapped data
            output_headers = mapped_headers
            parsed_rows = mapped_rows
            column_mapped = True

        except Exception as e:
            logger.warning(f"Column mapping failed, using raw data: {e}")
            # Fallback to raw data if mapping fails
            output_headers, parsed_rows = _parse_raw_data(headers, filtered_rows, errors)
    else:
        # No standard headers defined - use raw data
        output_headers, parsed_rows = _parse_raw_data(headers, filtered_rows, errors)

    # Clean headers for BigQuery column names
    cleaned_headers = [clean_header(h) for h in output_headers]

    # Ensure unique column names
    seen = {}
    unique_headers = []
    for h in cleaned_headers:
        if not h:
            h = "column"
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)

    # Convert row dicts to use cleaned headers
    final_rows = []
    for row in parsed_rows:
        final_row = {}
        for i, orig_header in enumerate(output_headers):
            if i < len(unique_headers):
                value = row.get(orig_header, "")
                final_row[unique_headers[i]] = parse_cell_value(value, unique_headers[i])
        final_rows.append(final_row)

    logger.info(
        f"Parsed {filename}: {len(final_rows)} rows, {skipped} skipped, mapped={column_mapped}",
        category=category.name,
        rows=len(final_rows),
        skipped=skipped,
        column_mapped=column_mapped,
    )

    return ParsedFile(
        filename=filename,
        headers=unique_headers,
        rows=final_rows,
        category=category,
        total_rows=len(raw_rows),
        skipped_rows=skipped,
        errors=errors,
        column_mapped=column_mapped,
        source_headers=source_headers,
    )


def _parse_raw_data(
    headers: List[str],
    rows: List[List[Any]],
    errors: List[str],
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse raw data without column mapping (fallback).

    Returns:
        Tuple of (headers, rows as list of dicts)
    """
    parsed_rows = []

    for row_idx, row in enumerate(rows):
        try:
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    row_dict[header] = row[col_idx]
                else:
                    row_dict[header] = ""
            parsed_rows.append(row_dict)
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    return headers, parsed_rows
